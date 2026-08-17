"""
soi — one reader for the SOI Corporation Complete Report balance sheets.

Both pipelines need these tables and previously read them separately:
  z pipeline     the LEVELS of land and inventories (weights for one year)
  flow pipeline  the DIFFERENCE in those levels across two years (acquisition)

Table 5.1 covers all corporations and Table 6.1 covers S corporations, so the
C-corporation figure is the difference.  Cells withheld for disclosure appear as
letter codes rather than numbers.  They are NOT read as zero: because the
C-corporation figure is a difference, a withheld S-corporation cell read as zero
overstates the C-corporation amount one for one.  They are estimated instead --
see _fill_suppressed -- using the published All Industries total to identify the
withheld residual.  The amounts are not negligible: in 2022 the withheld cells are
S-corporation land in utilities and real estate, together about $65 billion, or
roughly 8 percent of C-corporation land.
"""
from functools import lru_cache

import openpyxl
import pandas as pd

from params import check_schema, DATA_DIR, SECTOR_CODES, SECTOR_NAMES, SOI_BALANCE

# Major-industry header text -> canonical sector key.  SOI writes these headers
# in prose ("Agriculture, forestry, fishing, and hunting"), so match on a

check_schema(2, __name__)
# distinctive fragment rather than the full string.
_HEADER_KEY = {
    "agricultur": "11", "mining": "21", "utilit": "22", "construction": "23",
    "manufactur": "31-33", "wholesale": "42", "retail": "44-45",
    "transportation": "48-49", "information": "51", "finance": "52",
    "real estate": "53", "professional": "54", "management": "55",
    "administrative": "56", "educational": "61", "health care": "62",
    "arts": "71", "accommodation": "72", "other services": "81",
}
ITEMS = ("land", "inventories", "depreciable assets", "total assets")


@lru_cache(maxsize=None)
def _read_one(path, items=ITEMS):
    """Sum each item across the columns belonging to each sector ($ thousands).

    Returns (values, suppressed, all_industries).  SOI writes cells withheld for
    disclosure as a letter code rather than a number; those sectors are recorded
    in `suppressed` so the caller can estimate them, because reading them as zero
    would understate the item.  The All Industries column is published even when
    sector cells are not, so the total withheld is known.

    Cached on (path, items): both pipelines read the same workbooks, the flow
    pipeline reads two years, and the 2x2 reads each year twice.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=50, max_col=220, values_only=True))
    header = rows[4]
    out = {item: dict.fromkeys(SECTOR_CODES, 0.0) for item in items}
    sup = {item: set() for item in items}
    allind = {}
    for item in items:
        row = next((r for r in rows if r[0] and str(r[0]).strip().lower() == item), None)
        if row is None:
            continue
        allind[item] = float(row[1]) if isinstance(row[1], (int, float)) else None
        for j in range(1, len(header)):
            name = str(header[j]).strip().lower() if header[j] else ""
            if not name or name == "all industries":
                continue
            for frag, sector in _HEADER_KEY.items():
                if frag in name:
                    if isinstance(row[j], (int, float)):
                        out[item][sector] += float(row[j])
                    else:
                        sup[item].add(sector)
                    break
    return out, {k: frozenset(v) for k, v in sup.items()}, allind


def _fill_suppressed(scorp, sup, allind, allc):
    """Estimate S-corporation cells withheld for disclosure.

    The All Industries column gives the total, so the amount withheld is
    residual = All Industries - sum of everything published.  A sector can be
    partly published: SOI splits some sectors across more than one column, so a
    sector flagged as withheld may already carry a published amount.  The estimate
    is therefore ADDED to whatever is published for that sector, never substituted
    for it -- substituting discards the published part and makes the total fall
    below the published sum.  It is allocated across
    the withheld sectors in proportion to

        (S-corp total assets in sector) / (all-corp total assets in sector)
          x  (all-corp item in sector),

    then scaled so the estimates sum to the residual.  Allocating in proportion
    to the all-corporation distribution of the item alone would impose the same
    S-corporation share on every sector, which badly overstates sectors that are
    almost entirely C-corporation (utilities is the clearest case).  Reading the
    cells as zero is worse still: it understates the S-corporation subtraction and
    so overstates C-corporation holdings one for one.
    """
    # Total assets first: the estimator for every other item uses each sector's
    # S-corporation share of total assets, so that item must be complete before
    # the others are filled. Dict order otherwise follows the caller's request,
    # which puts total assets last.
    order = (["total assets"] + [k for k in sup if k != "total assets"])
    for item in order:
        sectors = sup.get(item, frozenset())
        if not sectors or allind.get(item) is None:
            continue
        residual = allind[item] - sum(scorp[item].values())
        if residual <= 0:
            continue
        weights = {}
        for s in sectors:
            assets_all = allc.get("total assets", {}).get(s, 0.0)
            assets_sc = scorp.get("total assets", {}).get(s, 0.0)
            share = (assets_sc / assets_all) if assets_all > 0 else 0.0
            weights[s] = max(share * allc.get(item, {}).get(s, 0.0), 0.0)
        total_w = sum(weights.values())
        for s in sectors:
            add = (residual * weights[s] / total_w if total_w > 0
                   else residual / len(sectors))
            scorp[item][s] += add
    return scorp


def clear_cache():
    """Drop cached workbook reads."""
    _read_one.cache_clear()


def levels(year, items=ITEMS, billions=True):
    """C-corporation balance-sheet levels by sector: Table 5.1 less Table 6.1.

    Returns a DataFrame indexed by sector display name, columns = items.
    """
    if year not in SOI_BALANCE:
        raise KeyError(f"no SOI balance sheet configured for {year}; "
                       f"have {sorted(SOI_BALANCE)}")
    f51, f61 = SOI_BALANCE[year]
    need = tuple(dict.fromkeys(tuple(items) + ("total assets",)))
    allc, _, _ = _read_one(DATA_DIR / f51, need)
    scorp_raw, sup, allind = _read_one(DATA_DIR / f61, need)
    # copy before filling: _read_one's return value is cached
    scorp = {k: dict(v) for k, v in scorp_raw.items()}
    scorp = _fill_suppressed(scorp, sup, allind, allc)
    scale = 1e6 if billions else 1.0        # $thousands -> $billions
    data = {item: [(allc[item][c] - scorp[item][c]) / scale for c in SECTOR_CODES]
            for item in items}
    return pd.DataFrame(data, index=[SECTOR_NAMES[c] for c in SECTOR_CODES])


def ccorp_share_of_depreciable(year):
    """C-corp share of all-corporation depreciable assets, by sector.

    Used to convert BEA corporate (C+S) produced-asset stocks to a C-corp basis.
    Assumes S corporations within an industry hold the same asset mix as C
    corporations, which is the approximation the paper discloses.
    """
    lv = levels(year, items=("depreciable assets",), billions=False)
    f51, f61 = SOI_BALANCE[year]
    allc, _, _ = _read_one(DATA_DIR / f51, ("depreciable assets", "total assets"))
    a = allc["depreciable assets"]
    return pd.Series(
        {SECTOR_NAMES[c]: (lv.loc[SECTOR_NAMES[c], "depreciable assets"] / a[c]
                           if a[c] > 0 else 1.0) for c in SECTOR_CODES})
