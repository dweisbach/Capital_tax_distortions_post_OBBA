"""
stocks — the industry-by-asset capital stock K_ia, C-corporation basis,
current-cost dollars.  Input to the z (primary) specification.

Three asset blocks, measured differently because they are published differently:

  equipment, structures, R&D   BEA detailed nonresidential net stocks, CURRENT
                               cost (the fixed-cost version must not be used for
                               composition shares).  All legal forms, so narrowed
                               to C corporations in two steps: BEA Table 4.1's
                               nine industry-group x asset-class corporate
                               shares, then the SOI C-corp share of depreciable
                               assets by industry.
  land, inventories            SOI balance sheets (Table 5.1 less Table 6.1).
                               Book value; land is rescaled to a current-cost
                               control -- see LAND_FACTOR in params.
  residential structures       BEA Table 5.1 corporate row, assigned to real
                               estate.  Small, and absent from the
                               nonresidential detail file.

Returns both the 12-category matrix and its 8-category collapse.
"""
import re
from functools import lru_cache

import openpyxl
import pandas as pd

from params import (check_schema, DATA_DIR, BEA_DETAIL, BEA_T41, BEA_T51, SECTOR_CODES,
                    SECTOR_NAMES, NAMES, BEA_NAICS_ALIAS, bea_asset_to_cat,
                    CATS_DETAILED, CAT8, DETAILED_TO_COMPARABLE, LAND_BASES,
                    LAND_PRIMARY, LAND_REFERENCE_YEAR)

check_schema(2, __name__)
import soi

_ASSET_CLASS = {
    "Equipment": "Equipment", "RD": "IPP", "Manufacturing_Struct": "Structures",
    "Communications_Struct": "Structures", "Farm_Struct": "Structures",
    "Commercial_HealthCare": "Structures", "Other_Nonres": "Structures",
    "OilGasMining_Struct": "Structures", "Power_Struct": "Structures",
    "Residential_Struct": "Structures",
}


def _bea_table(title_fragment, configured):
    """Locate a BEA Fixed Assets table by filename, else by its title row.

    BEA exports every interactive table as 'Table.csv', so a second download
    lands as 'Table (1).csv'.  Matching on the title row means the raw export
    names work without renaming.
    """
    p = DATA_DIR / configured
    if p.exists():
        return p
    for f in sorted(DATA_DIR.glob("*.csv")):
        try:
            head = f.open(encoding="utf-8", errors="ignore").readline()
        except OSError:
            continue
        if title_fragment.lower() in head.lower():
            return f
    raise FileNotFoundError(
        f"Could not find {configured!r} or any CSV whose title row names "
        f"{title_fragment!r} in {DATA_DIR}.")


def _bea_industry_map(wb):
    """BEA industry code -> canonical sector key, from the readme tab."""
    out = {}
    for title, code, _, naics in wb["readme"].iter_rows(
            min_row=15, max_row=130, max_col=4, values_only=True):
        if code and str(code).strip() and "----" not in str(code) \
                and str(code).strip() != "BEA CODE":
            m = re.match(r"\s*(\d{2})", str(naics))
            if m:
                two = m.group(1)
                out[str(code).strip()] = BEA_NAICS_ALIAS.get(two, two)
    return out


@lru_cache(maxsize=None)
def _bea_series():
    """Parse the BEA Datasets sheet ONCE and return (years, series).

    series is a tuple of (sector_name, category, values) with values aligned to
    years.  The sheet holds every industry x asset series for every vintage, and
    parsing it takes several seconds, so a run that needs two stock years should
    read it once rather than once per year.
    """
    wb = openpyxl.load_workbook(DATA_DIR / BEA_DETAIL, read_only=True, data_only=True)
    ind2sec = _bea_industry_map(wb)
    codes = sorted(ind2sec, key=len, reverse=True)
    ds = wb["Datasets"]
    header = list(next(ds.iter_rows(min_row=1, max_row=1, values_only=True)))
    years = tuple(int(re.search(r"(19|20)\d{2}", str(h)).group(0))
                  if re.search(r"(19|20)\d{2}", str(h)) else None for h in header)
    out = []
    for row in ds.iter_rows(min_row=2, values_only=True):
        label = row[0]
        if not label or not str(label).startswith("K1N"):
            continue
        rest = str(label)[3:].rsplit(".A", 1)[0]   # strip K1N prefix and .A suffix
        if len(rest) < 5:
            continue
        asset, ind = rest[-4:].upper(), rest[:-4]
        sector = next((ind2sec[c] for c in codes if ind.startswith(c)), None)
        if sector not in SECTOR_NAMES or asset in ("ST00", "IP00", "IPP0"):
            continue                              # roll-up totals
        cat = bea_asset_to_cat(asset)
        if cat:
            out.append((SECTOR_NAMES[sector], cat, row))
    return years, tuple(out)


def _produced_stocks(year):
    """All-legal-form produced-asset stocks by sector x category ($ billions)."""
    years, series = _bea_series()
    if year not in years:
        raise KeyError(f"BEA detail file has no column for {year}; "
                       f"it covers {min(y for y in years if y)}-"
                       f"{max(y for y in years if y)}")
    col = years.index(year)
    out = pd.DataFrame(0.0, index=NAMES, columns=CATS_DETAILED)
    for name, cat, row in series:
        value = row[col]
        if value is not None:
            out.loc[name, cat] += float(value) / 1e3      # $millions -> $billions
    return out


@lru_cache(maxsize=None)
def _corporate_shares(t41_year):
    """Nine corporate shares: three industry groups x three asset classes."""
    t = pd.read_csv(_bea_table("Table 4.1", BEA_T41), header=3, dtype=str)
    t.columns = [str(c).strip() for c in t.columns]

    def v(line):
        return float(t[t["Line"] == str(line)][str(t41_year)].iloc[0])

    lines = {"Farms": (6, 7, 8, 22, 23, 24),
             "Manufacturing": (10, 11, 12, 26, 27, 28),
             "Nonfarm": (14, 15, 16, 30, 31, 32)}
    return {g: {"Equipment": v(ce) / v(ae), "Structures": v(cs) / v(as_),
                "IPP": v(ci) / v(ai)}
            for g, (ae, as_, ai, ce, cs, ci) in lines.items()}


@lru_cache(maxsize=None)
def _corp_residential(year):
    """Corporate residential structures ($ billions), BEA Table 5.1."""
    t = pd.read_csv(_bea_table("Table 5.1", BEA_T51), header=3, dtype=str)
    t.columns = [str(c).strip() for c in t.columns]
    col = str(max(2017, year))                       # Table 5.1 begins in 2017
    row = t[t.iloc[:, 1].astype(str).str.strip() == "Corporate"]
    return float(row[col].iloc[0]) if len(row) else 0.0


@lru_cache(maxsize=None)
def land_factor(basis=LAND_PRIMARY):
    """Scale from SOI book land to the chosen current-value control.

    The denominator is the constructed book total in the reference year, read
    fresh, so the factor stays consistent with however land is currently
    measured.  One factor is applied to every year, so the land series does not
    move with the control, which is a residual and unstable year to year.
    """
    target = LAND_BASES[basis]
    if target is None:
        return 1.0                                    # SOI book value, unscaled
    book = soi.levels(LAND_REFERENCE_YEAR, items=("land",))["land"].sum()
    if book <= 0:
        raise ValueError("constructed SOI book land total is not positive")
    return target / book


def _group(sector):
    return "Farms" if sector == "11" else (
        "Manufacturing" if sector == "31-33" else "Nonfarm")


def clear_cache():
    """Drop cached file parses (call after editing an input file in a session)."""
    _bea_series.cache_clear()
    land_factor.cache_clear()
    _corporate_shares.cache_clear()
    _corp_residential.cache_clear()
    soi.clear_cache()


def build_K(year=2022, t41_year=None, land_basis=LAND_PRIMARY):
    """K_ia on a C-corporation basis ($ billions).  Returns (K12, K8).

    land_basis selects the current-value control for land (see params.LAND_BASES);
    the scale factor is derived from the constructed book total in the reference
    year, so the reference-year land stock equals the control exactly.
    """
    t41_year = t41_year or max(2017, year)           # Table 4.1 begins in 2017
    K = _produced_stocks(year)

    shares = _corporate_shares(t41_year)
    for sector in SECTOR_CODES:
        name = SECTOR_NAMES[sector]
        for cat, cls in _ASSET_CLASS.items():
            K.loc[name, cat] *= shares[_group(sector)][cls]

    cc_share = soi.ccorp_share_of_depreciable(year)
    produced = [c for c in CATS_DETAILED if c not in ("Land", "Inventories")]
    for name in NAMES:
        K.loc[name, produced] *= cc_share[name]

    lv = soi.levels(year, items=("land", "inventories"))
    K["Land"] = lv["land"] * land_factor(land_basis)
    K["Inventories"] = lv["inventories"]
    # All corporate residential is assigned to real estate, so the internally
    # consistent C-corp conversion is that sector's own share.
    re_name = SECTOR_NAMES["53"]
    K.loc[re_name, "Residential_Struct"] += _corp_residential(year) * cc_share[re_name]

    K8 = pd.DataFrame(0.0, index=NAMES, columns=CAT8)
    for c in CATS_DETAILED:
        K8[DETAILED_TO_COMPARABLE[c]] += K[c]
    return K[CATS_DETAILED], K8
